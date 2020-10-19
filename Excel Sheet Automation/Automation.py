import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path


def process_work_book(filename):
  
    wortk_book = xl.load_workbook(filename)
    sheet = wortk_book['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        corrected_price_heading = sheet.cell(1, 4)
        corrected_price_heading.value = 'Corrected_Price'

    value = Reference(
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4
    )

    chart = BarChart()
    chart.add_data(value)
    sheet.add_chart(chart, 'e2')
    wortk_book.save(filename)

path = Path()
new_work_book = path.glob('*.xlsx')
needed_files = []

for file in new_work_book:
    needed_files.append(file)

print(needed_files)
process_work_book(needed_files[0])


